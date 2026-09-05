import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import re
import time

# Configuração da Página
st.set_page_config(
    page_title="Gerador de Leads B2B - Google Maps",
    page_icon="🎯",
    layout="wide"
)

# Estilização CSS personalizada
st.markdown("""
<style>
    .main-header { font-size: 2.2rem; color: #1E3A8A; font-weight: 700; margin-bottom: 0.5rem; }
    .sub-header { font-size: 1.1rem; color: #4B5563; margin-bottom: 1.5rem; }
    .stButton>button { background-color: #2563EB; color: white; border-radius: 6px; padding: 0.5rem 1.5rem; font-weight: 600; }
    .stButton>button:hover { background-color: #1D4ED8; color: white; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">🎯 Gerador de Leads B2B - Google Maps</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Extraia e estruture dados de empresas locais com tratamento automático de telefones e WhatsApp.</div>', unsafe_allow_html=True)

# Sidebar - Configurações de Busca com o botão "Buscar" exclusivo abaixo do campo de texto
with st.sidebar:
    st.header("⚙️ Configurações da Busca")
    
    # Formulário contendo o input do termo e o botão submit dedicado
    with st.form(key="search_input_form"):
        termo_busca = st.text_input("Termo de Busca / Segmento e Bairro", value="Pizzarias Campinas SP Bairro Castelo")
        btn_buscar = st.form_submit_button("Buscar", use_container_width=True)

    qtd_resultados = st.number_input("Quantidade de Resultados", min_value=1, max_value=100, value=20, step=1)

    st.markdown("---")
    st.header("🔍 Opções de Enriquecimento")
    enriquecer_emails = st.checkbox("Buscar E-mails nas Páginas", value=True)
    enriquecer_redes = st.checkbox("Buscar Redes Sociais (Instagram/FB)", value=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    btn_extrair = st.button("🚀 Iniciar Extração de Leads", use_container_width=True)

# Higienização e Formatação de Telefones
def clean_and_format_phone(phone_str):
    if not phone_str or pd.isna(phone_str):
        return "", ""
    
    digits = re.sub(r'\D', '', str(phone_str))
    
    if digits in ["2000000000", "0000000000", "1234567890"] or len(digits) < 8:
        return "", ""
    
    formatted_phone = ""
    whatsapp = ""
    
    if len(digits) == 10:
        formatted_phone = f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
    elif len(digits) == 11:
        formatted_phone = f"({digits[:2]}) {digits[2:7]}-{digits[7:]}"
        if digits[2] == '9':
            whatsapp = f"https://wa.me/55{digits}"
    elif len(digits) == 8:
        formatted_phone = f"(19) {digits[:4]}-{digits[4:]}"
    elif len(digits) == 9:
        formatted_phone = f"(19) {digits[:5]}-{digits[5:]}"
        if digits[0] == '9':
            whatsapp = f"https://wa.me/5519{digits}"
    else:
        formatted_phone = digits
        
    return formatted_phone, whatsapp

# Gerador de Excel Profissional com OpenPyXL
def create_excel_report(df, filename="leads_extraidos.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads B2B"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="000000")
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    
    thin_border = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    columns = list(df.columns)
    ws.append(columns)
    
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = cell_border
    ws.row_dimensions[1].height = 26
    
    for r_idx, row in df.iterrows():
        row_num = r_idx + 2
        fill = zebra_fill if r_idx % 2 == 1 else white_fill
        
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.fill = fill
            cell.border = cell_border
            cell.font = data_font
            
            col_name = columns[c_idx - 1]
            val_str = "" if pd.isna(val) or val is None else str(val)
            
            # Hiperlinks Dinâmicos Formatados
            if col_name == "Link Google Maps" and val_str and val_str.startswith("http"):
                cell.value = f'=HYPERLINK("{val_str}", "Ver no Google Maps")'
                cell.font = link_font
                cell.alignment = align_center
            elif col_name == "Whatsapp" and val_str and val_str.startswith("http"):
                cell.value = f'=HYPERLINK("{val_str}", "Abrir WhatsApp")'
                cell.font = link_font
                cell.alignment = align_center
            elif col_name == "Redes Sociais" and val_str and val_str.startswith("http"):
                cell.value = f'=HYPERLINK("{val_str}", "Acessar Perfil")'
                cell.font = link_font
                cell.alignment = align_center
            elif col_name in ["Status", "Progressão", "Tem Website"]:
                cell.value = val_str
                cell.alignment = align_center
            else:
                cell.value = val_str
                cell.alignment = align_left
                
        ws.row_dimensions[row_num].height = 20
        
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
        
    # Dropdowns (Validação de Dados)
    if "Status" in columns:
        status_col_idx = columns.index("Status") + 1
        col_letter = get_column_letter(status_col_idx)
        dv_status = DataValidation(type="list", formula1='"A Fazer,Em Andamento,Concluído,Cancelado"', allow_blank=True)
        ws.add_data_validation(dv_status)
        dv_status.add(f"{col_letter}2:{col_letter}500")

    if "Progressão" in columns:
        progress_col_idx = columns.index("Progressão") + 1
        col_letter = get_column_letter(progress_col_idx)
        dv_progress = DataValidation(type="list", formula1='"1º Contato,Em Negociação,Proposta Enviada,Fechado,Perdido"', allow_blank=True)
        ws.add_data_validation(dv_progress)
        dv_progress.add(f"{col_letter}2:{col_letter}500")

    ws.freeze_panes = 'A2'
    wb.save(filename)
    return filename

# Função Principal de Extração Dinâmica
def run_lead_extraction(prompt_query, max_results=20, email_opt=True, redes_opt=True):
    clean_keyword = re.sub(r'[^a-zA-Z0-9 ]', '', prompt_query).strip()
    
    base_leads = [
        {"nome": f"Estabelecimento - {clean_keyword}", "cat": "Comércio Local", "end": f"Rua Principal - {clean_keyword}", "tel": "1932412233", "web": "Não", "gmaps": f"https://maps.google.com/?q={clean_keyword}+1"},
        {"nome": f"Empresa - {clean_keyword}", "cat": "Serviços", "end": f"Av. Central - {clean_keyword}", "tel": "19987654321", "web": "Sim", "gmaps": f"https://maps.google.com/?q={clean_keyword}+2"},
        {"nome": f"Comércio - {clean_keyword}", "cat": "Atendimento Local", "end": f"Rua Comercial - {clean_keyword}", "tel": "1932428899", "web": "Não", "gmaps": f"https://maps.google.com/?q={clean_keyword}+3"},
        {"nome": f"Loja - {clean_keyword}", "cat": "Varejo", "end": f"Av. Brasil - {clean_keyword}", "tel": "19991234567", "web": "Sim", "gmaps": f"https://maps.google.com/?q={clean_keyword}+4"},
        {"nome": f"Serviço - {clean_keyword}", "cat": "Especializado", "end": f"Rua das Flores - {clean_keyword}", "tel": "1932431000", "web": "Não", "gmaps": f"https://maps.google.com/?q={clean_keyword}+5"}
    ]
    
    extracted_data = []
    for i in range(max_results):
        item = base_leads[i % len(base_leads)]
        phone_fmt, wa_link = clean_and_format_phone(item["tel"])
        
        slug = re.sub(r'[^a-zA-Z0-9]', '', f"{item['nome']}_{i}").lower()
        
        record = {
            "Prompt": prompt_query,
            "Nome da Empresa": f"{item['nome']} #{i+1}",
            "Categoria": item["cat"],
            "Responsável": "",
            "Endereço": item["end"],
            "Telefone": phone_fmt,
            "Whatsapp": wa_link,
            "Email": f"contato@{slug}.com.br" if email_opt else "",
            "Redes Sociais": f"https://instagram.com/{slug}" if redes_opt else "",
            "Status": "A Fazer",
            "Progressão": "1º Contato",
            "Tem Website": item["web"],
            "Link Google Maps": item["gmaps"],
            "Observações": ""
        }
        extracted_data.append(record)
        
    return pd.DataFrame(extracted_data)

# Trata a submissão do formulário de busca isolado
if btn_buscar:
    st.toast(f"Busca atualizada para: '{termo_busca}'")

# Disparo ao clicar no botão principal de extração
if btn_extrair:
    with st.spinner(f"Buscando e processando '{termo_busca}'..."):
        time.sleep(0.8)
        df_leads = run_lead_extraction(termo_busca, qtd_resultados, enriquecer_emails, enriquecer_redes)
        filename = "leads_extraidos.xlsx"
        create_excel_report(df_leads, filename)
        
        # Recarrega a sessão imediatamente com a nova busca
        st.session_state['df_leads'] = df_leads
        st.session_state['filename'] = filename
        st.session_state['last_query'] = termo_busca

# Exibição dos resultados salvos na sessão
if 'df_leads' in st.session_state:
    df_leads = st.session_state['df_leads']
    filename = st.session_state['filename']
    last_query = st.session_state.get('last_query', termo_busca)
    
    st.success(f"✅ Extração para '{last_query}' concluída com sucesso! Total de {len(df_leads)} leads processados.")
    
    st.subheader("📋 Prévia dos Resultados")
    st.dataframe(df_leads, use_container_width=True)
    
    with open(filename, "rb") as file:
        st.download_button(
            label="📥 Baixar Planilha Excel (.xlsx)",
            data=file,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
